# -*- coding: utf-8 -*-
"""
File Name: init_femb.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 7/15/2016 11:47:39 AM
Last modified: 10/18/2016 4:37:37 PM
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
import numpy as np
#import scipy as sp
#import pylab as pl
from scipy.integrate import simps
import openpyxl as px
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Font, Side, PatternFill
import numpy as np
import struct
import os
from scripts.detect_peaks import detect_peaks
import matplotlib
#matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy import stats
import sys
import glob
import pickle
import re
from scripts.Data_Analysis import Data_Analysis
import glob
import pickle
import warnings
import gc
from user_settings import user_editable_settings
settings = user_editable_settings()
analyze = Data_Analysis()

####################################################################################################################################################
#This function should be run after the RMS file with the slope data has already been made.  It takes in the root folder, the specific directory
#to look through, the existing workbook to save to, and the total number of chips.  If will read the binary file created for RMS data analysis.
#It will find the mean and RMS of the data, and convert it to other units, and format it all for final analysis in the RMS spreadsheet.
class Energy:
    def Energy(self, cal_directory, data_directory, samples, analysis = "peak"):
        
        self.events_directory = data_directory + "_Events\\"
        try: 
            os.makedirs(self.events_directory)
        except OSError:
            if os.path.exists(self.events_directory):
                pass
        #Get array of all the baselines, peaks and integral gains
        RMS_filename = cal_directory + "Internal_RMS_Data.xlsx"
        wb = px.load_workbook(RMS_filename) 
        sheet_title = "25.0,3.0,200"
        print (sheet_title)
        ws = wb[sheet_title]
        baselines = [[],[],[],[]]
        integral_ratio = [[],[],[],[]]
        peak_gain = [[],[],[],[]]
        for chip in range(settings.chip_num):
            for chn in range(16):
                baselines[chip].append(ws.cell(row=4 + chip, column=2 + chn).value)
                integral_ratio[chip].append(ws.cell(row=32 + chip, column=2 + chn).value)
                peak_gain[chip].append(ws.cell(row=25 + chip, column=2 + chn).value)
    
        all_event_energies = []
        event_summary = []
        #Go through each event number and make sure a packet for all 4 chips exists (maybe not necessary?)
        file_rec = self.events_directory + "Analysis_Summary.txt"
        screendisplay = sys.stdout
        sys.stdout = open(file_rec, "w+")
        for i in range(samples):
            files = []
            for j in range(settings.chip_num):
                files.append(data_directory + "\\Chip{}_Packet{}.dat".format(j,i))
            exists = 0
            for file1 in files:
                if os.path.isfile(file1):
                    exists = exists + 1
            if (exists > settings.chip_num -1):
                self.event_directory = self.events_directory + "Event_{}_".format(i)
                energies = []
                #Because Chip 2 isn't ready yet
                for chip in [0,1,3]:
                    #Get waveform for each channel.  Returns the chip data (subtracted by the baseline, so it should
                    #sit at 0 ADC counts), and the index of which channels had peaks detected
                    self.event = i
                    self.chip = chip
                    returned = self.UnpackData(data = files[chip], base = baselines[chip], chip_num = chip)
                    if (returned == None):
                        continue
#                    fig = returned[2]
                    chip_data = returned[1]
                    index = returned[0]
                    energy_map = dict()
                    #Loop through channels that had peaks detected
                    for chn in index:
                        filename = self.event_directory + "Chip{}_Chn{}_Energy".format(chip, chn)+".jpg"
                        peaks = detect_peaks(x=chip_data[chn], mph=self.threshold, mpd=30)
                        #Make sure only one peak is detected for that channel
                        if (len(peaks) > 1):
                            print ("Energy_Analysis --> There are {} peaks in Event {}, Chip {}, Chn{}".format(len(peaks), i, chip, chn))
                            pass
                        else:
                            #Return the bounds that you want to integrate within, as well as the colors to plot it,
                            #so you know how they were found.  Sometimes, I've found then even when subtracting the 
                            #nominal baseline, some channels are at a higher baseline, which throws off the integration
                            #calculation, so it also returns the last baseline correction number, if necessary.
                            calibration_bounds, colors, subtraction = self.find_bounds(chip_data[chn], peaks[0])
                            
                            #If there was something to subtract, do it.  Check to see if any peaks are actually registered
                            #because some might have only been found because of this issue.  If there really aren't any
                            #peaks, then exit out of this analysis
                            if (subtraction != None):
                                print ("Event {}, Chip {}, Channel {} had to subtract {}".format(i, chip, chn, subtraction))
                                for index in range(len(chip_data[chn])):
                                    chip_data[chn][index] = chip_data[chn][index] - subtraction
                                peaks = detect_peaks(x=chip_data[chn], mph=self.threshold, mpd=30)
                                if (peaks < 1):
                                    break
                                
                            #Make the plot of the single pulse being integrated
                            if (self.to_print == True):
                                fig,ax = self.quickPlot(data = chip_data[chn], name = filename)
                                ax.set_xlim([-1,71])
                                ax.set_title("Event{}_Chip{}_Channel{}".format(self.event, self.chip, chn), fontsize=30)
                                ax.scatter(peaks[0]/2.0,chip_data[chn][peaks[0]],color='r',marker='x',lw=2, s=25)
                                for num,xc in enumerate(calibration_bounds):
                                    ax.axvline(x=xc/2.0, color=colors[num], linestyle='--')
                                

                            if (analysis == "peak"):
                                peak_height = chip_data[chn][peaks[0]]
                                charge = peak_height / peak_gain[chip][chn]

                            #Use the integral method for all samples within the bounds
                            elif (analysis == "area"):
                                integral_window = chip_data[chn][calibration_bounds[0]:calibration_bounds[1]]
                                total_sum = 0
                                for point in integral_window:
                                    total_sum = total_sum + point
                                charge = total_sum / integral_ratio[chip][chn]

                            #Find the energy (check if this is ok?)
                            energy = charge / (4.1 / 570)
                            energies.append(energy)
                            energy_map[str(chn)] = energy
                            
                            #Finish off the single channel plot with the energy details
                            if (self.to_print == True):
                                subtitle_text = (.25,.94)
                                ax.annotate("Sum = {:.2f} bins\nCharge = {:.2f} fC\nEnergy = {:.2f} keV".format(
                                    total_sum, charge, energy), xy=subtitle_text,  xycoords='axes fraction',
                                    xytext=subtitle_text, textcoords='axes fraction',
                                    horizontalalignment='center', verticalalignment='center', fontsize = 20
                                    )
                                ticks = [tick for tick in ax.get_yticks() if tick >=0]
                                ax.set_yticks(ticks)
                                fig.savefig (filename)
                                plt.close(fig)
                            
                    #With all of the channels in a chip analyzed, plot the overlay of all of them
                    self.plotAll(data = chip_data, ref = energy_map, 
                                 filename = self.event_directory + "Chip{}".format(chip) + "overlayed.jpg")
                            
                #Find the total energy of an event and save it for the histogram and the text file output
                sum_of_energies = sum(energies)
                if (sum_of_energies > 0):
                    all_event_energies.append(sum_of_energies)
                    event_summary.append([sum_of_energies, i])

        #Put the text file in order of energies, so you can quickly look up events for a certain energy result
        file_rec = self.events_directory + "Energy_Summary.txt"
        screendisplay = sys.stdout
        sys.stdout = open(file_rec, "w+")
        print ("Energy, Event")
        for i in sorted(event_summary):
            print ("Energy: {:.2f} keV,Event: {}".format(i[0], i[1]))
        sys.stdout.close()
        sys.stdout = screendisplay
        return all_event_energies
    
    def UnpackData(self, data, base, chip_num = 0):

        fileinfo  = os.stat(data)
        filelength = fileinfo.st_size
    
        with open(data, 'rb') as f:
            raw_data = f.read(filelength)
            f.close()
        FACE_check = struct.unpack_from(">%dH"%(self.BPS + 8),raw_data)

        face_index = -1
        for i in range (self.BPS):
            if (FACE_check[i] == 0xFACE):
                face_index = i
                break
            
        if (face_index == -1):
            print ("FACE not detected")
            print_tuple = []
            for i in range(len(FACE_check)):
                print_tuple.append(hex(FACE_check[i]))
            
            print (print_tuple)
            return
        
        #print ("FACE detected at " + str(face_index))
        
        shorts = filelength/2
        full_data = struct.unpack_from(">%dH"%shorts,raw_data)
        full_data = full_data[face_index:]
        test_length = len(full_data[face_index:])
        full_samples = test_length // self.BPS
        
        ch0 = []
        ch1 = []
        ch2 = []
        ch3 = []
        ch4 = []
        ch5 = []
        ch6 = []
        ch7 = []
        ch8 = []
        ch9 = []
        ch10 = []
        ch11 = []
        ch12 = []
        ch13 = []
        ch14 = []
        ch15 = []

        #Correct for the baseline right away
        for i in range (full_samples):
            ch7.append((full_data[(self.BPS*i)+1] & 0x0FFF) - base[7])
            ch6.append((((full_data[(self.BPS*i)+2] & 0x00FF) << 4) + ((full_data[(self.BPS*i)+1] & 0xF000) >> 12)) - base[6])
            ch5.append((((full_data[(self.BPS*i)+3] & 0x000F) << 8) + ((full_data[(self.BPS*i)+2] & 0xFF00) >> 8)) - base[5])
            ch4.append((((full_data[(self.BPS*i)+3] & 0xFFF0) >> 4)) - base[4])
            ch3.append((full_data[(self.BPS*i)+4] & 0x0FFF) - base[3])
            ch2.append((((full_data[(self.BPS*i)+5] & 0x00FF) << 4) + ((full_data[(self.BPS*i)+4] & 0xF000) >> 12)) - base[2])
            ch1.append((((full_data[(self.BPS*i)+6] & 0x000F) << 8) + ((full_data[(self.BPS*i)+5] & 0xFF00) >> 8)) - base[1])
            ch0.append((((full_data[(self.BPS*i)+6] & 0xFFF0) >> 4)) - base[0])
            ch15.append((full_data[(self.BPS*i)+7] & 0x0FFF) - base[15])
            ch14.append((((full_data[(self.BPS*i)+8] & 0x00FF) << 4) + ((full_data[(self.BPS*i)+7] & 0xF000) >> 12)) - base[14])
            ch13.append((((full_data[(self.BPS*i)+9] & 0x000F) << 8) + ((full_data[(self.BPS*i)+8] & 0xFF00) >> 8)) - base[13])
            ch12.append((((full_data[(self.BPS*i)+9] & 0xFFF0) >> 4)) - base[12])
            ch11.append((full_data[(self.BPS*i)+10] & 0x0FFF) - base[11])
            ch10.append((((full_data[(self.BPS*i)+11] & 0x00FF) << 4) + ((full_data[(self.BPS*i)+10] & 0xF000) >> 12)) - base[10])
            ch9.append((((full_data[(self.BPS*i)+12] & 0x000F) << 8) + ((full_data[(self.BPS*i)+11] & 0xFF00) >> 8)) - base[9])
            ch8.append((((full_data[(self.BPS*i)+12] & 0xFFF0) >> 4)) - base[8])
            
        chip = [ch0,ch1,ch2,ch3,ch4,ch5,ch6,ch7,ch8,ch9,ch10,ch11,ch12,ch13,ch14,ch15]

        peaks_index = []
        for i in range(16):
            peaks = detect_peaks(x=chip[i], mph = self.threshold, mpd=10)
            peaks_index.append(np.array(peaks).tolist())
            
#If a peak is near a stuck bit, give it a different color
#We could also ignore any events with multiple peaks over a threshold, but that means
#we would ignore Comptons, right?
        peak_marker = [[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
        for i in range(len(peaks_index)):
            for j in range(len(peaks_index[i])):
                index = peaks_index[i][j]
                mod = chip[i][index] % 64
                if ((mod == 0) or (mod == 1) or (mod == 2) or (mod == 63) or (mod == 62)):
                    peak_marker[i].append('g')
                else:
                    peak_marker[i].append('r')
#            for k in sorted(to_delete, reverse = True):
#                del peaks_index[i][k]

#If no peaks are found, don't bother plotting or trying to do anything else
        if (all( not chn_peaks for chn_peaks in peaks_index )):
            return None
            
            
        if (len(ch7) == len(ch8)):
            all_equal = True
        else:
            all_equal = False
            
        if (all_equal == True):
            length = len(ch7)
        else:
            print("Energy Analysis --> Not all channels in {} have the same length!".format(data))
            
        index = []
        if (self.to_print != True):
            for i in range(16):
                if (peaks_index[i]):
                    index.append(i)
                    
            return [index, chip]
            
        time = []
        
        for i in range(length):
            time.append(0.5 * i)
            
        fig = plt.figure(figsize=(16, 12), dpi=80)
        overlay_ax = fig.add_subplot(1,1,1)
        overlay_ax.spines['top'].set_color('none')
        overlay_ax.spines['bottom'].set_color('none')
        overlay_ax.spines['left'].set_color('none')
        overlay_ax.spines['right'].set_color('none')
        overlay_ax.tick_params(labelcolor='w', top='off', bottom='off', left='off', right='off')
        overlay_ax.set_xlabel('Time (us)', fontsize = 24)
        overlay_ax.set_ylabel('ADC Counts', fontsize = 24)
        overlay_ax.yaxis.set_label_coords(-0.015,0.5)
        overlay_ax.set_title("Event{}_Chip{}".format(self.event, self.chip), fontsize=30)
        
        ax1 = fig.add_subplot(16,1,16)
        ax1.plot(time, ch0)
#        ax1.plot([0,1], [base[0]] * 2, color = 'r')
#        ax1.axhline(y=base[0], color='r', linestyle='--')
        #Don't ask me why, but if this isn't run, the main X-axis gets lost and can't be brought back.
        fig.canvas.draw()
        ax1.set_xticklabels(ax1.get_xticklabels(), fontsize=12)
        if (peaks_index[0]):
            index.append(0)
            for j in range(len(peaks_index[0])):
                x_coord = peaks_index[0][j]
                ax1.scatter([x_coord/2.0],[chip[0][x_coord]],color=peak_marker[0][j],marker='x',lw=2, s=25)
        ax1.set_title("Channel 0")
        ax2 = ax1.twinx()
        ax2.set_ylabel("Chn 0", rotation = 0)
        ax2.spines['top'].set_color('none')
        ax2.spines['bottom'].set_color('none')
        ax2.spines['left'].set_color('none')
        ax2.spines['right'].set_color('none')
        ax2.tick_params(labelcolor='w', top='off', bottom='off', left='off', right='off')
        
        #Make sure each subplot has reasonable y limits
        limits = list(ax1.get_ylim())
        if (limits[0] > self.plot_bounds[0]):
            limits[0] = self.plot_bounds[0]

        if (limits[1] < self.plot_bounds[1]):
            limits[1] = self.plot_bounds[1]
        ax1.set_ylim(limits)
        ticks = [tick for tick in ax1.get_yticks() if tick >=0]
        ax1.set_yticks(ticks)
        fig.canvas.draw()
        for i in range(1,16,1):
            ax = fig.add_subplot(16,1,16-i, sharex=ax1)
            ax.plot(time, chip[i])
            #Don't ask me why, but if this isn't run, the main X-axis gets lost and can't be brought back.
            fig.canvas.draw()
#            ax.plot([0,1], [base[i]] * 2, color = 'r')
            ax.set_xticklabels(ax.get_xticklabels(), visible=False)
            if (peaks_index[i]):
                index.append(i)
                for j in range(len(peaks_index[i])):
                    x_coord = peaks_index[i][j]
                    ax.scatter([x_coord/2.0],[chip[i][x_coord]],color=peak_marker[i][j],marker='x',lw=2, s=25)
            ax2 = ax.twinx()
            ax2.set_ylabel("Chn " + str(i), rotation = 0)
    #            ax2.spines['top'].set_color('none')
    #            ax2.spines['bottom'].set_color('none')
    #            ax2.spines['left'].set_color('none')
    #            ax2.spines['right'].set_color('none')
            ax2.tick_params(labelcolor='w', top='off', bottom='off', left='off', right='off')
            pos1 = ax2.get_position() # get the original position 
            pos2 = [pos1.x0 + 0.025, pos1.y0 + 0.005,  pos1.width , pos1.height ] 
            ax2.set_position(pos2) # set a new position
            
            limits = list(ax.get_ylim())
            if (limits[0] > self.plot_bounds[0]):
                limits[0] = self.plot_bounds[0]
    
            if (limits[1] < self.plot_bounds[1]):
                limits[1] = self.plot_bounds[1]
            ax.set_ylim(limits)
            ticks = [tick for tick in ax.get_yticks() if tick >=0]
            ax.set_yticks(ticks)
    
        plt.subplots_adjust(wspace=0, hspace=0, top = 0.95, bottom = 0.05, right = 0.95, left = 0.05)
        fig.savefig (self.event_directory + "Chip{}".format(chip_num)+".jpg")
        plt.close(fig)
        return [index, chip, fig]
    
    def quickPlot(self, data, name):
        time = []
        for i in range(len(data)):
            time.append(0.5 * i)
        fig = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig.add_subplot(1,1,1)
        ax.set_xlabel('Time (us)', fontsize = 24)
        ax.set_ylabel('ADC Counts', fontsize = 24)
        plt.plot(time, data)
        
#        fig.savefig (name)
        plt.close(fig)
        return fig,ax
        
    #Receives the data for 16 channels in a chip, and an index of what the channels used for energy analysis were
    #and what their energies were
    def plotAll(self, data, ref, filename):
        summary = dict()
        time = []
        for i in range(len(data[0])):
            time.append(0.5 * i)
        fig = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig.add_subplot(1,1,1)
        ax.set_xlabel('Time (us)', fontsize = 24)
        ax.set_ylabel('ADC Counts', fontsize = 24)
        ax.set_title("Event{}_Chip{}".format(self.event, self.chip), fontsize=30)
        for chn in range(len(data)):
            #If it was used for energy analysis, take a note of which channel was assigned which color
            if (str(chn) in ref):
                ret = ax.plot(time, data[chn])
                c = ret[-1].get_color()
                summary[str(chn)] = [ref[(str(chn))],c]
            #If not, make sure it wasn't part of the 'bad baselines' as described below.  If those are printed,
            #they confuse the plot.  Note that they will print if they have a pulse associated with them.
            elif (chn not in self.bad_channels[self.chip]):
                ax.plot(time, data[chn], color='black')
                
        #Negative ticks for ADC counts are confusing
        ticks = [tick for tick in ax.get_yticks() if tick >=0]
        ax.set_yticks(ticks)
        total_energy = []

        #Print an energy summary for each channel, in the color it was printed
        last_key = 0
        for i, key in enumerate(summary):
            ax.text(0.02, 0.97 - (0.03 * i), "Chn {}: Energy = {:.2f} keV".format(key, summary[key][0]),
            verticalalignment='bottom', horizontalalignment='left',
            transform=ax.transAxes,
            color=summary[key][1], fontsize=18)
            total_energy.append(summary[key][0])
            last_key = i
           
        #If more than one channel is printed, also print the energy sum
        if (last_key > 0):
            ax.text(0.02, 0.97 - (0.03 * (last_key + 1)), "Total Energy = {:.2f} keV".format(sum(total_energy)),
                verticalalignment='bottom', horizontalalignment='left',
                transform=ax.transAxes,
                color='black', fontsize=18)
        fig.savefig (filename)
        plt.close(fig)
    
    #Given a peak and the full data, find reasonable bounds for integration.  What it will do is use a window
    #as defined in the class self variables to analyze what the RMS is.  If it's too high, that means we're still 
    #on the slope of the pulse.  So we move left or right until the RMS is low enough that we must be back at the baseline
    #A red bound will mean it correctly found a good stable spot.  Green means that it went too far and I told the
    #process to just stop.  Blue means we hit the edge of the data, maybe the pulse was too close to the beginning
    #or end.
    def find_bounds(self, corrected_chip_data, peak):
        #find lower bound
        window = []
        index = 5
        rms = self.rms_check + 1
        color_l = 'r'
        while (rms > self.rms_check):
            index = index + 1
            point_to_analyze = peak - index
            rms_window = corrected_chip_data[point_to_analyze - self.rms_bound: point_to_analyze + self.rms_bound]
            rms = np.mean(rms_window)
            if (point_to_analyze < (peak - self.calibration_bounds[0])):
                color_l = 'g'
                break
            if ((point_to_analyze - self.rms_bound) < 1):
                color_l = 'b'
                break
        window.append(point_to_analyze)
        #find upper bound
        index = 5
        rms = self.rms_check + 1
        color_h = 'r'
        while (rms > self.rms_check):
            index = index + 1
            point_to_analyze = peak + index
            rms_window = corrected_chip_data[point_to_analyze - self.rms_bound: point_to_analyze + self.rms_bound]
            rms = np.mean(rms_window)
            if (point_to_analyze > (peak + self.calibration_bounds[1])):
                color_h = 'g'
                break
            if ((point_to_analyze - self.rms_bound) > (len(corrected_chip_data) - 1)):
                color_h = 'b'
                break
        window.append(point_to_analyze)
        
        outside_bounds = []
        for sample in (corrected_chip_data[0:window[0]]):
            outside_bounds.append(sample)
        for sample in (corrected_chip_data[window[1]:]):
            outside_bounds.append(sample)
            
        mean_outside_bounds = np.mean(outside_bounds)
        subtraction = None
        if (abs(mean_outside_bounds) > self.bad_offset):
            subtraction = mean_outside_bounds
        
        return window,[color_l,color_h], subtraction
    
    def fixData(self, data):
        for i in range(len(data)):
            datum = data[i]
            mod = datum % 64
            if ((mod == 0) or (mod == 1) or (mod == 2) or (mod == 63) or (mod == 62)):
                
                if(i == 0):
                    data[i] = data[i+1]
#                    print ("{} became {}".format(data[i], data[i+1]))
                elif (i > (len(data) -2)):
                    data[i] = data[i-1]
#                    print ("{} became {}".format(data[i], data[i-1]))
                else:
                    data[i] = (data[i-1] + data[i+1])/2
#                    print ("{} became {}".format(data[i], (data[i-1] + data[i+1])/2))
                    
        return data
    
    def __init__(self):
        self.BPS = 13 #Bytes per sample.  The one for "0xFACE" and then 12 bytes for 16 channels at 12 bits each.
        self.channels = 16
        self.threshold = 30     #~6 sigma above typical RMS noise
        self.start_of_packet = (b"\xde\xad\xbe\xef")
        self.start_of_sample = (b"\xfa\xce")
        self.debug_file_name = "\Data_Analysis_Debug.txt"
        self.screendisplay = None
        self.notice1_every = 10000
        self.notice2_every = 50000
        self.to_print = True
        self.calibration_bounds = [40,40]
        self.plot_bounds = [-20, 80]
        self.rms_check = 5
        self.rms_bound = 5
        self.bad_channels = [[9,13],[11],[],[9,11,13,15]]
        self.bad_offset = 5